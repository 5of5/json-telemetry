#!/usr/bin/env python3
"""Generate per-binary docs from the workbook + 25 REFINEMENT map mixers.

The 25 names are sheet 05_MAP_COVERAGE_25 (sealed registry). These are not
invented map types (B5/B10). They remix already-processed JSON telemetry
into one map's declared neighborhood. Source data is never rewritten.
"""
from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
XLSX = ROOT / "TRACN Binary Repository v1 (1).xlsx"
CATALOG = Path(__file__).resolve().parent / "catalog" / "operators.json"
DOC_DIR = Path(__file__).resolve().parent / "catalog" / "binaries"
MAP_DIR = Path(__file__).resolve().parent / "maps"
OUT_OPS = ROOT / "crates" / "operators"

FAMILY_KIND = {
    "COMPANY": ["Company"],
    "COMPETITOR": ["Company"],
    "PEOPLE": ["Person"],
    "BUYER": ["Person", "Account"],
    "SELLER": ["Company", "Product"],
    "SYNDICATE": ["Investor", "Syndicate"],
    "PARTNER": ["Company"],
    "VENTURE_CAPITAL": ["Investor", "Fund"],
    "EVENT": ["Event", "MarketSignal"],
    "PRODUCT": ["Product"],
    "CUSTOMER": ["Customer", "Account"],
    "MARKET": ["Market", "Category"],
    "ACCOUNT": ["Account"],
    "SIGNAL": ["MarketSignal"],
    "CLAIM": ["Claim"],
    "SOURCE": ["Source"],
    "PORTFOLIO": ["Portfolio", "MarketMap", "Workspace"],
    "CONTENT": ["Content"],
    "MARKET_MAP": ["MarketMap", "MapNode", "Segment"],
}


def slug_map(name: str) -> str:
    s = re.sub(r"[^A-Za-z0-9]+", "_", name.strip()).strip("_").upper()
    return s


def split_list(raw) -> list[str]:
    if raw is None:
        return []
    s = str(raw).strip()
    if s in {"", "—", "-", "n/a"}:
        return []
    parts = re.split(r"[|,]", s)
    out = []
    for p in parts:
        p = re.sub(r"\s+", " ", p).strip()
        if p and p not in {"—", "-"}:
            out.append(p)
    return out


def sheet_rows(wb, name: str, header_token: str) -> list[dict]:
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    hdr = start = None
    for i, r in enumerate(rows):
        if r and r[0] == header_token:
            hdr, start = [str(c) if c is not None else "" for c in r], i + 1
            break
    if hdr is None:
        raise SystemExit(f"no header {header_token!r} in {name}")
    out = []
    for r in rows[start:]:
        if not r or r[0] is None:
            continue
        d = dict(zip(hdr, list(r) + [None] * (len(hdr) - len(r))))
        out.append(d)
    return out


def why_from_01(wb) -> dict[str, dict]:
    out = {}
    for d in sheet_rows(wb, "01_BINARY_CATALOG", "binary_id"):
        bid = str(d.get("binary_id") or "")
        if bid.startswith("BIN."):
            out[bid] = d
    return out


def maps_05_13(wb) -> list[dict]:
    cov = sheet_rows(wb, "05_MAP_COVERAGE_25", "#")
    lang = {str(d.get("map_type")): d for d in sheet_rows(wb, "13_MAP_LANGUAGE_25", "#")}
    maps = []
    for d in cov:
        try:
            n = int(float(d["#"]))
        except (TypeError, ValueError):
            continue
        name = str(d["map_type"]).strip()
        L = lang.get(name, {})
        maps.append(
            {
                "n": n,
                "map_type": name,
                "slug": slug_map(name),
                "intent": d.get("intent") or "",
                "family": d.get("family") or "",
                "category_group": d.get("category_group") or "",
                "primary_binaries": split_list(d.get("primary_binaries")),
                "required_anchor_tags": split_list(d.get("required_anchor_tags")),
                "required_rel_types": split_list(d.get("required_rel_types")),
                "required_signal_types": split_list(d.get("required_signal_types")),
                "min_ent": d.get("min_ent"),
                "budget": d.get("budget"),
                "depth": d.get("depth"),
                "map_language_tags": split_list(L.get("map_language_tags")),
                "cast_rule": L.get("cast_rule") or "",
                "semantic_increase": L.get("semantic_increase") or "",
            }
        )
    maps.sort(key=lambda m: m["n"])
    if len(maps) != 25:
        raise SystemExit(f"expected 25 maps, got {len(maps)}")
    return maps


def refiner_spec(m: dict) -> dict:
    node_types: list[str] = []
    for fam in m["primary_binaries"]:
        for k in FAMILY_KIND.get(fam, []):
            if k not in node_types:
                node_types.append(k)
    slug = m["slug"]
    crate = f"aria_ref_{slug.lower()}"
    tags = list(dict.fromkeys(m["required_anchor_tags"] + [m["map_type"], "MAP_TYPE"]))
    return {
        "anchor_tags": tags,
        "binary_id": f"BIN.REF.{slug}",
        "class": "REFINEMENT",
        "crate": crate,
        "default_limit": None,
        "layer": "REFINEMENT",
        "neo4j_pass": True,
        "node_types": node_types,
        "operator": f"REF.{slug}",
        "package": crate.replace("_", "-"),
        "parent": "MARKET_MAP",
        "pass_through": False,
        "property_key": None,
        "relationship_types": m["required_rel_types"],
        "result_definition_ref": f"map.ref.{slug.lower()}",
        "retrieval_step": "memory",
        "taxonomy": m["map_type"],
        "telemetry_fork": f"aria/ref/{slug.lower()}@v1",
        "verify": True,
        "wave": None,
    }


CARGO_TOML = """# Generated. Unique operator crate; AriA is linked via aria-operator.
[package]
name = {name!r}
description = {desc!r}
version.workspace = true
edition.workspace = true
license.workspace = true
authors.workspace = true
repository.workspace = true
homepage.workspace = true
documentation.workspace = true
rust-version.workspace = true
publish = false

[lib]
path = "src/lib.rs"

[[bin]]
name = {name!r}
path = "src/main.rs"

[dependencies]
aria-json-telemetry = {{ workspace = true, features = ["cli"] }}

[lints]
workspace = true
"""

LIB_RS = """//! Unique operator {binary_id} ({operator}).
//!
//! Map mixer: slices already-tagged JSON telemetry into this sealed
//! market-map type. Source bytes are never rewritten.

/// Catalog identity.
pub const BINARY_ID: &str = {binary_id};
/// Operator name on the closed envelope.
pub const OPERATOR: &str = {operator};
/// This crate's frozen spec (sheet row).
pub const SPEC: &str = include_str!("../spec.json");

/// Run this operator on a host JSON payload.
pub fn run(payload: &[u8]) -> Result<aria_operator::OperatorEnvelope, aria_operator::OperatorError> {{
    aria_operator::run_spec(SPEC, payload, &aria_operator::RunOpts::default())
}}
"""

MAIN_RS = """fn main() {{
    std::process::exit(aria_operator::bin_main({pkg}::SPEC));
}}
"""


def write_crate(spec: dict) -> None:
    pkg = spec["package"]
    dest = OUT_OPS / pkg
    dest.mkdir(parents=True, exist_ok=True)
    (dest / "src").mkdir(exist_ok=True)
    (dest / "spec.json").write_text(json.dumps(spec, indent=2, sort_keys=True) + "\n")
    desc = f"Aria map mixer {spec['binary_id']} ({spec['taxonomy']})"
    (dest / "Cargo.toml").write_text(CARGO_TOML.format(name=pkg, desc=desc))
    ident = pkg.replace("-", "_")
    (dest / "src" / "lib.rs").write_text(
        LIB_RS.format(
            binary_id=json.dumps(spec["binary_id"]),
            operator=json.dumps(spec["operator"]),
        )
    )
    (dest / "src" / "main.rs").write_text(MAIN_RS.format(pkg=ident))


def binary_doc(spec: dict, why: dict, maps: list[dict], consumed_by: list[str]) -> str:
    bid = spec["binary_id"]
    row = why.get(bid, {})
    why_txt = (row.get("why_this_belongs_to_Aria_MAX_anchor") or "").strip()
    notes = (row.get("notes_from_prior_sheet") or "").strip()
    layer = spec.get("layer") or ""
    cls = spec.get("class") or ""
    parent = spec.get("parent") or "—"
    op = spec.get("operator") or ""
    use = f"`work --binary {bid}` or `work --json` with `ops: [\"{bid}\"]`"
    if layer == "REFINEMENT":
        fn = (
            "Map mixer. Ingests the same JSON any operator ingests "
            "(raw graph, or already-processed `aria-work-v1` callback). "
            "Returns ONLY the neighborhood this sealed map type is allowed "
            "to consume (sheet 05 kinds/rels/tags). Missing data is omitted. "
            "Source bytes are not rewritten. Mode 2 graphics consume this "
            "envelope; AriA does not Judge."
        )
        why_txt = why_txt or (
            f"Sealed map type {spec.get('taxonomy')} from sheet 05. "
            "One dump × 25 mixers is the viral coefficient: the same tagged "
            "telemetry fans out into 25 structured map JSON results without "
            "a second Trust write."
        )
    elif layer == "HOST":
        fn = "Host toolkit. Not a research operator. Empty limitation, no Φ (B6)."
    elif layer == "TRANSFORM":
        fn = "AriA transformer pass-through of the ingested graph. Never Judge. Never Trust."
    elif cls == "TAG":
        fn = (
            "Tag operator. Does not name a new entity. Tags or reads an existing "
            "node. Role-tags (BUYER/COMPETITOR/…) require the specific tag on the record."
        )
    elif cls == "REL":
        fn = "Relationship residual. Returns only this rel type and the endpoints actually used."
    elif cls == "PROP":
        fn = "Property residual. Echoes one declared key when present. Never mints Trust."
    elif cls == "NODE":
        fn = "Node residual. Returns only this kind. Empty is no-finding, not a guess."
    else:
        fn = (
            "Closed operator. Returns only its declared kinds as structured JSON. "
            "Independent of other binaries' scores (B2)."
        )
    lines = [
        f"# {bid}",
        "",
        f"**Operator:** `{op}` · **layer:** {layer} · **class:** {cls} · **parent:** {parent}",
        f"**Crate:** `{spec.get('package')}` · **verify:** {spec.get('verify')} · **result:** `{spec.get('result_definition_ref')}`",
        "",
        "## Why",
        "",
        why_txt or "Listed Binary Repository v1 identity. Independent calculation; no borrowed scores.",
        "",
        "## Function",
        "",
        fn,
        "",
        "## Use",
        "",
        use,
        "",
        "Production callback: working vertical or nothing. Forget is not delete — original payload stays in `source`.",
        "",
        "## Declared neighborhood",
        "",
        f"- node types: {', '.join(spec.get('node_types') or []) or '—'}",
        f"- relationships: {', '.join(spec.get('relationship_types') or []) or '—'}",
        f"- anchor tags: {', '.join(spec.get('anchor_tags') or []) or '—'}",
        f"- property key: {spec.get('property_key') or '—'}",
        "",
        "## Maps that consume this",
        "",
    ]
    if consumed_by:
        for m in consumed_by:
            lines.append(f"- {m}")
    else:
        lines.append("- (not a primary binary on sheet 05, or is itself a map mixer)")
    if notes:
        lines += ["", "## Sheet notes", "", notes]
    lines += [
        "",
        "## Important",
        "",
        "- No Trust / Use / Goal fields.",
        "- Does not rewrite other binaries' verticals.",
        "- Empty declared types → omitted from `aria-work-v1` results.",
        "",
    ]
    return "\n".join(lines)


def main() -> None:
    wb = load_workbook(XLSX, data_only=True, read_only=True)
    maps = maps_05_13(wb)
    why = why_from_01(wb)
    wb.close()

    catalog: list[dict] = json.loads(CATALOG.read_text())
    existing = {s["binary_id"] for s in catalog}
    refiners = [refiner_spec(m) for m in maps]
    added = [s for s in refiners if s["binary_id"] not in existing]
    catalog.extend(added)
    catalog.sort(key=lambda s: s["binary_id"])
    ids = [s["binary_id"] for s in catalog]
    if len(ids) != len(set(ids)):
        raise SystemExit("duplicate binary_id")
    CATALOG.write_text(json.dumps(catalog, indent=2, sort_keys=True) + "\n")
    for s in added:
        write_crate(s)

    consume: dict[str, list[str]] = {s["binary_id"]: [] for s in catalog}
    fam_to_bin = {
        "COMPANY": "BIN.COMPANY",
        "COMPETITOR": "BIN.COMPETITOR",
        "PEOPLE": "BIN.PEOPLE",
        "BUYER": "BIN.BUYER",
        "SELLER": "BIN.SELLER",
        "SYNDICATE": "BIN.SYNDICATE",
        "PARTNER": "BIN.PARTNER",
        "VENTURE_CAPITAL": "BIN.VENTURE_CAPITAL",
        "EVENT": "BIN.EVENT",
        "PRODUCT": "BIN.PRODUCT",
        "CUSTOMER": "BIN.CUSTOMER",
        "MARKET": "BIN.MARKET",
        "ACCOUNT": "BIN.ACCOUNT",
        "SIGNAL": "BIN.SIGNAL",
        "CLAIM": "BIN.CLAIM",
        "SOURCE": "BIN.SOURCE",
        "PORTFOLIO": "BIN.PORTFOLIO",
        "MARKET_MAP": "BIN.MARKET_MAP",
    }
    for m in maps:
        ref_id = f"BIN.REF.{m['slug']}"
        for fam in m["primary_binaries"]:
            bid = fam_to_bin.get(fam)
            if bid and bid in consume:
                consume[bid].append(f"{m['n']:02d} {m['map_type']} → {ref_id}")

    DOC_DIR.mkdir(parents=True, exist_ok=True)
    MAP_DIR.mkdir(parents=True, exist_ok=True)
    by_id = {s["binary_id"]: s for s in catalog}
    for spec in catalog:
        p = DOC_DIR / f"{spec['binary_id'].replace('.', '_')}.md"
        p.write_text(binary_doc(spec, why, maps, consume.get(spec["binary_id"], [])))

    index = ["# Binary identities", "", f"{len(catalog)} closed operators. Generated from Binary Repository v1 (xlsx 01/05/11/13/14).", "", "## Index", ""]
    for spec in catalog:
        fname = spec["binary_id"].replace(".", "_") + ".md"
        index.append(f"- [`{spec['binary_id']}`](binaries/{fname}) — {spec['layer']}/{spec['class']} `{spec['operator']}`")
    (Path(__file__).resolve().parent / "catalog" / "INDEX.md").write_text("\n".join(index) + "\n")

    maps_md = [
        "# 25 sealed market-map mixers",
        "",
        "Sheet `05_MAP_COVERAGE_25` + `13_MAP_LANGUAGE_25`. These are **refinement**",
        "operators (`BIN.REF.*`). They run on already-tagged JSON telemetry",
        "(a dump callback, a worker `aria-work-v1`, or a raw graph that already",
        "carries kinds/rels/tags). They do **not** invent entities, scores, or Trust.",
        "The original dump remains the evidence set; each mixer is a standalone",
        "map-shaped slice of that same data.",
        "",
        "Viral coefficient: one processed payload × 25 mixers = 25 structured map JSON",
        "results. Mode 2 enrichment reads these envelopes to draw the graphic.",
        "AriA still never Judges.",
        "",
        "| # | Map | `BIN.REF.*` | intent | family | primary binaries | rels |",
        "|---|---|---|---|---|---|---|",
    ]
    for m in maps:
        maps_md.append(
            f"| {m['n']} | {m['map_type']} | `BIN.REF.{m['slug']}` | {m['intent']} | {m['family']} | "
            f"{', '.join(m['primary_binaries'])} | {', '.join(m['required_rel_types'])} |"
        )
    maps_md += ["", "## Per map", ""]
    for m in maps:
        spec = by_id[f"BIN.REF.{m['slug']}"]
        maps_md += [
            f"### {m['n']:02d}. {m['map_type']}",
            "",
            f"- **binary:** `{spec['binary_id']}`",
            f"- **use:** `work --binary {spec['binary_id']}` on a dump callback or tagged graph",
            f"- **intent / family / group:** {m['intent']} / {m['family']} / {m['category_group']}",
            f"- **kinds:** {', '.join(spec['node_types'])}",
            f"- **rels:** {', '.join(spec['relationship_types'])}",
            f"- **cast rule (13):** {m['cast_rule']}",
            f"- **language tags:** {', '.join(m['map_language_tags'][:12])}{'…' if len(m['map_language_tags'])>12 else ''}",
            "",
            (m["semantic_increase"] or "").strip(),
            "",
        ]
    (MAP_DIR / "MAPS.md").write_text("\n".join(maps_md) + "\n")
    print(f"catalog {len(catalog)} (added {len(added)} refiners)")
    print(f"docs {len(list(DOC_DIR.glob('*.md')))} -> {DOC_DIR}")
    print(f"maps -> {MAP_DIR / 'MAPS.md'}")


if __name__ == "__main__":
    main()
